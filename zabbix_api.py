import requests
import openpyxl

# 创建Excel工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# Zabbix API相关信息
zabbix_api_url = 'http://**.**.**.**/api_jsonrpc.php'
zabbix_username = 'Admin'
zabbix_password = 'zabbix'


# 登录Zabbix获取访问令牌
login_payload = {
    'jsonrpc': '2.0',
    'method': 'user.login',
    'params': {
        'user': zabbix_username,
        'password': zabbix_password
    },
    'id': 1
}

response = requests.post(zabbix_api_url, json=login_payload)
response_data = response.json()

if 'result' in response_data:
    auth_token = response_data['result']
    print('登录成功，令牌：', auth_token)

    # 获取"ceni-bigdata"主机群组的ID
    get_group_payload = {
        'jsonrpc': '2.0',
        'method': 'hostgroup.get',
        'params': {
            'output': ['groupid'],
            'filter': {
                'name': ['ceni-bigdata']
            }
        },
        'auth': auth_token,
        'id': 2
    }

    response = requests.post(zabbix_api_url, json=get_group_payload)
    response_data = response.json()

    if 'result' in response_data:
        group_id = response_data['result'][0]['groupid']

        # 获取"ceni-bigdata"主机群组下所有主机的ID和名称
        get_hosts_payload = {
            'jsonrpc': '2.0',
            'method': 'host.get',
            'params': {
                'output': ['hostid', 'name'],
                'groupids': [group_id]
            },
            'auth': auth_token,
            'id': 3
        }

        response = requests.post(zabbix_api_url, json=get_hosts_payload)
        response_data = response.json()

        if 'result' in response_data:
            hosts = response_data['result']
            
            # 按照主机名进行排序
            hosts.sort(key=lambda x: int(x['name'].split('-')[-1]))

            # 写入主机名
            column = 1
            for host in hosts:
                host_name = host['name']
                sheet.cell(row=1, column=column, value=host_name)
                column += 1

            # 获取主机数目
            num_hosts = len(hosts)

            # 获取主机的CPU占用、内存占用和根分区占用
            for index, host in enumerate(hosts):
                host_id = host['hostid']
                host_name = host['name']

                # 获取主机的CPU占用
                get_cpu_payload = {
                    'jsonrpc': '2.0',
                    'method': 'item.get',
                    'params': {
                        'output': ['lastvalue'],
                        'hostids': [host_id],
                        'search': {
                            'key_': 'system.cpu.util[,user]'
                        },
                        'sortfield': 'name'
                    },
                    'auth': auth_token,
                    'id': 4
                }

                response = requests.post(zabbix_api_url, json=get_cpu_payload)
                response_data = response.json()

                if 'result' in response_data:
                    cpu_usage = response_data['result'][0]['lastvalue']
                    sheet.cell(row=2, column=index+1, value=float(cpu_usage))
                    sheet.cell(row=2, column=index+1, value=f"{float(cpu_usage):.2f}%")

                # 获取主机的内存占用
                get_memory_payload = {
                    'jsonrpc': '2.0',
                    'method': 'item.get',
                    'params': {
                        'output': ['lastvalue'],
                        'hostids': [host_id],
                        'search': {
                            'key_': 'vm.memory.size[pavailable]'
                        },
                        'sortfield': 'name'
                    },
                    'auth': auth_token,
                    'id': 5
                }

                response = requests.post(zabbix_api_url, json=get_memory_payload)
                response_data = response.json()

                if 'result' in response_data:
                    memory_usage = response_data['result'][0]['lastvalue']
                    memory_usage_percentage = float(memory_usage)
                    inverted_percentage = 100 - memory_usage_percentage
                    sheet.cell(row=3, column=index+1, value=f"{inverted_percentage:.2f}%")
               # 获取主机的根分区占用
                get_disk_payload = {
                    'jsonrpc': '2.0',
                    'method': 'item.get',
                    'params': {
                        'output': ['lastvalue'],
                        'hostids': [host_id],
                        'search': {
                            'key_': 'vfs.fs.size[/,pused]'
                        },
                        'sortfield': 'name'
                    },
                    'auth': auth_token,
                    'id': 6
                }

                response = requests.post(zabbix_api_url, json=get_disk_payload)
                response_data = response.json()

                if 'result' in response_data:
                    disk_usage = response_data['result'][0]['lastvalue']
                    sheet.cell(row=4, column=index+1, value=float(disk_usage))
                    sheet.cell(row=4, column=index+1, value=f"{float(disk_usage):.2f}%")

        else:
            print('未找到主机。')

    else:
        print('获取主机群组失败。')

else:
    print('登录失败。')

# 保存Excel文件
workbook.save('host_stats.xlsx')
print('主机统计信息已保存到host_stats.xlsx文件。')
